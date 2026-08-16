#!/usr/bin/env python3
"""
INFORME xlsx -> informe.json, ready for scripts/import-informe.ts.

The workbook is a weekly sales report kept by hand, so it is parsed defensively:

* Every weekday sheet has a DIFFERENT column layout (LUNES starts at column B,
  JUEVES has no FECHA column, ...), so the header row is located per sheet.
* The operator does not use the columns consistently — the unit price is
  sometimes typed into the UNIDAD column. TOTAL is the one figure always
  present and correct, so it arbitrates: whichever reading of
  (KGS, UNIDAD, P/UNIT) reproduces TOTAL is the true one.
* Whether a line is fabric (Kg) or combos (Units) is therefore decided by the
  ARITHMETIC, not by the wording of the description.
* Product names carry typos, varying word order, widths ("56 CM") and the
  occasional real lot number ("LOT. 7650"). Colors normalize through an
  explicit alias table so "CELEZTE", "CELESTE" and "AZUL CELESTE" converge.

Nothing here writes to a database — it only produces JSON to review.
"""
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

# Week of INFORME 001. The sheets carry "FECHA: 19.01.26" etc.
DAYS = {
    "LUNES": "2026-01-19", "MARTES": "2026-01-20", "MIERCOLES": "2026-01-21",
    "JUEVES": "2026-01-22", "VIERNES": "2026-01-23", "SABADO": "2026-01-24",
}
# Column indexes per sheet, read off each sheet's own header row.
LAYOUT = {
    "LUNES":     dict(n0=1, cli=3, prod=4, kg=5, ud=6, pu=7, tot=8, transf=9, efect=10, bs=11, obs=13),
    "MARTES":    dict(n0=0, cli=2, prod=3, kg=4, ud=5, pu=6, tot=7, efect=8, bs=9, obs=11),
    "MIERCOLES": dict(n0=0, cli=2, prod=3, kg=4, ud=5, pu=6, tot=7, efect=8, bs=9, obs=12),
    # JUEVES/VIERNES head their cash column "USD" (not "EFECTIVO") — it was
    # missed on the first pass and every DIVISAS sale of those two days parsed
    # as unpaid, which the VIERNES CSV re-sent by the client exposed.
    "JUEVES":    dict(n0=0, cli=1, prod=2, kg=3, ud=4, pu=5, tot=6, bs=7, efect=8, obs=9),
    "VIERNES":   dict(n0=1, cli=2, prod=3, kg=4, ud=5, pu=6, tot=7, bs=8, efect=9, obs=10),
    "SABADO":    dict(n0=0, cli=1, prod=2, kg=3, ud=4, pu=5, tot=6, efect=7, bs=8, obs=11),
}
# "TASA BCV" as written on each sheet. SABADO has none (no trading) — carry Friday's.
RATES = {"LUNES": 344.5, "MARTES": 344.5, "MIERCOLES": 347.26,
         "JUEVES": 349.92, "VIERNES": 353.0, "SABADO": 353.0}

# ---- product vocabulary -----------------------------------------------------

# Order matters twice over. "DRY FIT" must beat "FIT", and the counted articles
# (Combo/Cuellos/Puños) must beat the fabric words: "Combos Pique Blanco" is a
# COMBO, and if it were filed as fabricType "Piqué" its batch _id would collide
# with the ROLL batch of the same colour+NM — ingressStock rejects a batch that
# already exists as another productType, so this is a hard failure, not a typo.
FABRICS = [
    (("COMBOS", "COMBO"), "Combo"),
    (("CUELLOS", "CUELLO"), "Cuellos"),
    (("PUÑOS", "PUNOS"), "Puños"),
    (("DRY FIT", "DRYFIT"), "Dry Fit"),
    (("INTERLOCK",), "Interlock"),
    (("JERSEY",), "Jersey"),
    # "Chemise y piqué es lo mismo" — the client's own correction (casilla 11).
    # Without it a future INFORME creates a second catalogue of the same cloth.
    (("PIQUET", "PIQUE", "PIQU", "CHEMISE", "CHEMIS"), "Piqué"),
    (("RIBB", "RIB", "IBB"), "Ribb"),
    (("TELA DE SEGUNDA", "SEGUNDA"), "Tela de segunda"),
]
# Sold off a roll, by weight — never by the piece.
WEIGHT_FABRICS = {"Jersey", "Piqué", "Ribb", "Interlock", "Dry Fit", "Tela de segunda"}
COUNTED_FABRICS = {"Combo", "Cuellos", "Puños"}

# Typos and short forms seen in INFORME 001, mapped to one spelling each.
COLOR_ALIASES = {
    "BLANCOS": "Blanco", "BLANCO": "Blanco", "BLACNO": "Blanco",
    "NEGRO": "Negro", "NEGROS": "Negro",
    "MARINO": "Azul Marino", "AZUL MARINO": "Azul Marino", "AZULMARINO": "Azul Marino",
    "REY": "Azul Rey", "AZUL REY": "Azul Rey",
    "CELESTE": "Azul Celeste", "CELEZTE": "Azul Celeste", "AZUL CELESTE": "Azul Celeste",
    "ROJO": "Rojo", "AMARILLO": "Amarillo", "NARANJA": "Naranja",
    "TURQUEZA": "Turquesa", "TURQUESA": "Turquesa",
    "LILA": "Lila", "BEIGE": "Beige", "MELANGE": "Melange",
    "GRIS RATON": "Gris Ratón", "GRIS RARON": "Gris Ratón", "GRIS": "Gris",
    "VERDE PERICO": "Verde Perico", "VERDE MANZANA": "Verde Manzana",
    "VERDE BOTELLA": "Verde Botella", "VERDE": "Verde",
    "NAZARENO": "Nazareno", "UNIPOLO BLANCO": "Blanco",
    "COLORES VARIADOS": "Variado", "VARIADO": "Variado", "VARIADOS": "Variado",
}
# Words that describe the cut, not the article.
NOISE = re.compile(
    r"\b(ABIERTO|ABIERTA|TUBULAR|\d{2,3}\s*CM|LOT\.?\s*\d+|\(.*?\)|DE|EL|LA)\b", re.I)
NM_RE = re.compile(r"\b(\d{2})\s*/\s*1\b|\b(20|24|30|26|28)1\b")
LOT_RE = re.compile(r"LOT\.?\s*(\d+)", re.I)


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", ".").strip())
        except ValueError:
            return None
    return None


def txt(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def close(a, b, tol=0.02):
    return a is not None and b is not None and abs(a - b) <= max(tol, abs(b) * 0.01)


def resolve_line(kg, ud, pu, tot):
    """(qty, unit, price) — TOTAL arbitrates which column is really the price."""
    for qty, unit in ((kg, "Kg"), (ud, "Units")):
        if qty and pu and close(qty * pu, tot):
            return qty, unit, pu
    if kg and ud and not pu and close(kg * ud, tot):      # price typed into UNIDAD
        return kg, "Kg", ud
    for qty, unit in ((kg, "Kg"), (ud, "Units")):
        if qty and tot and qty > 0:
            return qty, unit, round(tot / qty, 4)         # derive from the total
    if kg and pu:
        return kg, "Kg", pu
    if ud and pu:
        return ud, "Units", pu
    return None


def parse_product(desc: str, unit: str):
    """description + resolved unit -> (color, nm, fabricType, productType, unit, lot|None)."""
    up = strip_accents(desc).upper()
    lot = LOT_RE.search(up)
    lot = lot.group(1) if lot else None

    fabric = None
    for keys, label in FABRICS:
        for k in keys:
            if k in up:
                fabric = label
                up = up.replace(k, " ")
                break
        if fabric:
            break

    # The wording decides, not the column the operator happened to use:
    #  * anything named "combos"/"cuellos" is counted, even when the count was
    #    typed into KGS ("COMBOS | KGS 3 | P/U 1 | TOTAL 3");
    #  * a fabric sold by weight is in Kg even when the quantity was typed into
    #    UNIDAD — "RIBB AZUL REY | UD 1 | 7.9 | 7.9" is the ~1 kg companion
    #    piece that goes out with a jersey, not one countable unit.
    if fabric in COUNTED_FABRICS:
        unit, product_type = "Units", "COMBO"
    elif fabric in WEIGHT_FABRICS:
        unit, product_type = "Kg", "ROLL"
    else:
        product_type = "COMBO" if unit == "Units" else "ROLL"
        fabric = fabric or ("Combo" if product_type == "COMBO" else "Tela")

    m = NM_RE.search(up)
    nm = f"{m.group(1) or m.group(2)}/1" if m else None
    if m:
        up = up[:m.start()] + " " + up[m.end():]

    up = NOISE.sub(" ", up)
    rest = re.sub(r"[^A-ZÑ ]", " ", up)
    rest = re.sub(r"\s+", " ", rest).strip()

    color = None
    for alias in sorted(COLOR_ALIASES, key=len, reverse=True):   # longest first
        if re.search(rf"\b{re.escape(alias)}\b", rest):
            color = COLOR_ALIASES[alias]
            break
    if color is None:
        color = rest.title() if rest else "Variado"

    return color, nm, fabric, product_type, unit, lot


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "docs/spread/INFORME 001.xlsx")
    wb = openpyxl.load_workbook(src, data_only=True)

    sales, unresolved = [], []
    for day, date in DAYS.items():
        if day not in wb.sheetnames:
            continue
        L, ws, cur = LAYOUT[day], wb[day], None
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
            def g(k):
                return row[L[k]] if k in L and L[k] < len(row) else None
            n0, prod = txt(g("n0")), txt(g("prod"))
            kg, ud, pu, tot = num(g("kg")), num(g("ud")), num(g("pu")), num(g("tot"))

            if re.fullmatch(r"\d{3,6}", n0):
                if cur and cur["lines"]:
                    sales.append(cur)
                cur = dict(ref=n0, day=day, date=date, client=txt(g("cli")),
                           rate=RATES[day], lines=[],
                           paidUsdCash=0.0, paidUsdTransfer=0.0, paidBs=0.0, obs="")
            if cur is None:
                continue

            if prod:
                r = resolve_line(kg, ud, pu, tot)
                if not r:
                    if tot:
                        unresolved.append((day, cur["ref"], prod, kg, ud, pu, tot))
                    continue
                qty, unit, price = r
                color, nm, fabric, ptype, unit, lot = parse_product(prod, unit)
                # parse_product may correct the unit (see its comment); the
                # money is fixed, so re-derive the price from the total.
                if ptype == "COMBO" and unit == "Units":
                    qty = float(round(qty))
                cur["lines"].append(dict(raw=prod, qty=qty, unit=unit, price=price,
                                         color=color, nm=nm, fabricType=fabric,
                                         productType=ptype, lot=lot))
            else:
                # Subtotal / payment row: only the money columns matter here.
                for key, field in (("transf", "paidUsdTransfer"), ("efect", "paidUsdCash"), ("bs", "paidBs")):
                    v = num(g(key)) if key in L else None
                    if v:
                        cur[field] += v
                o = txt(g("obs"))
                if o and not cur["obs"]:
                    cur["obs"] = o
        if cur and cur["lines"]:
            sales.append(cur)

    # A missing NM is filled with the week's dominant count rather than left blank:
    # nm is part of the batch _id, and "sin NM" would fragment the catalogue.
    counts = Counter(l["nm"] for s in sales for l in s["lines"] if l["nm"])
    default_nm = counts.most_common(1)[0][0] if counts else "20/1"
    invented_nm = 0
    for s in sales:
        for l in s["lines"]:
            if not l["nm"]:
                l["nm"], l["nmInvented"] = default_nm, True
                invented_nm += 1

    out = dict(
        source=src.name,
        week=sorted({s["date"] for s in sales}),
        defaultNm=default_nm,
        sales=sales,
        stats=dict(sales=len(sales),
                   lines=sum(len(s["lines"]) for s in sales),
                   clients=len({s["client"] for s in sales if s["client"]}),
                   unresolvedLines=len(unresolved),
                   inventedNm=invented_nm,
                   realLots=sum(1 for s in sales for l in s["lines"] if l["lot"])),
    )
    dest = src.parent / "informe.json"
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf8")
    print(json.dumps(out["stats"], indent=1))
    print("wrote", dest)
    if unresolved:
        print("\nUNRESOLVED LINES:")
        for u in unresolved:
            print("  ", u)


if __name__ == "__main__":
    main()
